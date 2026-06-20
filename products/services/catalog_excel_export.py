"""Экспорт каталога товаров и разделов в Excel (.xlsx)."""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook

from products.models import Category, Product

PRODUCT_HEADERS = (
    "Код 1С (id)",
    "Артикул",
    "Название",
    "Путь раздела",
    "Корневой раздел",
    "Код категории",
    "Категория",
    "Розничная цена",
    "Оптовая цена",
    "Остаток",
    "Ед. изм.",
    "Активен",
    "Обновлён в БД",
)

CATEGORY_HEADERS = (
    "Код категории",
    "Название",
    "Код родителя",
    "Родитель",
    "Путь раздела",
    "Корневой раздел",
    "Активна",
    "Товаров в разделе",
)


def _category_maps(categories: list[Category]) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    by_id = {str(c.pk): c for c in categories}
    paths: dict[str, str] = {}
    roots: dict[str, str] = {}

    def _path_for(cat_id: str) -> str:
        if cat_id in paths:
            return paths[cat_id]
        cat = by_id.get(cat_id)
        if not cat:
            paths[cat_id] = ""
            return ""
        parent_id = str(cat.parent_id) if cat.parent_id else ""
        if parent_id and parent_id in by_id:
            parent_path = _path_for(parent_id)
            full = f"{parent_path} / {cat.name}" if parent_path else cat.name
        else:
            full = cat.name
        paths[cat_id] = full
        return full

    for cat_id in by_id:
        _path_for(cat_id)
        full = paths.get(cat_id, "")
        roots[cat_id] = full.split(" / ")[0] if full else ""

    parent_names = {
        str(c.pk): (by_id[str(c.parent_id)].name if c.parent_id and str(c.parent_id) in by_id else "")
        for c in categories
    }
    return paths, roots, parent_names


def _product_row(p: Product, *, paths: dict[str, str], roots: dict[str, str]) -> list:
    cat = p.category
    cat_id = str(cat.pk)
    upd = p.updated_at
    if upd is not None and timezone.is_aware(upd):
        upd = timezone.localtime(upd).replace(tzinfo=None)
    elif upd is not None:
        upd = upd.replace(tzinfo=None)
    return [
        str(p.pk),
        (p.sku or "").strip(),
        p.name,
        paths.get(cat_id, cat.name),
        roots.get(cat_id, cat.name),
        cat_id,
        cat.name,
        float(p.retail_price),
        float(p.wholesale_price),
        int(p.stock),
        (p.unit or "").strip(),
        bool(p.is_active),
        upd,
    ]


def build_catalog_excel_bytes(*, active_only: bool = False) -> bytes:
    """
    XLSX с двумя листами:
    - «Товары» — вся номенклатура с полным путём раздела;
    - «Разделы» — дерево категорий и число товаров.
    """
    categories = list(Category.objects.all().order_by("name"))
    paths, roots, parent_names = _category_maps(categories)

    product_counts: dict[str, int] = defaultdict(int)
    qs = Product.objects.select_related("category").order_by("category__name", "name")
    if active_only:
        qs = qs.filter(is_active=True)
    products = list(qs)
    for p in products:
        product_counts[str(p.category_id)] += 1

    buf = BytesIO()
    wb = Workbook(write_only=True)

    ws_products = wb.create_sheet(title="Товары", index=0)
    ws_products.append(list(PRODUCT_HEADERS))
    for p in products:
        ws_products.append(_product_row(p, paths=paths, roots=roots))

    ws_categories = wb.create_sheet(title="Разделы", index=1)
    ws_categories.append(list(CATEGORY_HEADERS))
    for cat in categories:
        cat_id = str(cat.pk)
        parent_id = str(cat.parent_id) if cat.parent_id else ""
        ws_categories.append(
            [
                cat_id,
                cat.name,
                parent_id,
                parent_names.get(cat_id, ""),
                paths.get(cat_id, cat.name),
                roots.get(cat_id, cat.name),
                bool(cat.is_active),
                int(product_counts.get(cat_id, 0)),
            ]
        )

    wb.save(buf)
    return buf.getvalue()
